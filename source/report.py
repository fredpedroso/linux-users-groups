#!/usr/bin/env python3
# -*- coding: utf-8 -*-

""" Report module. """

import time
import shutil
import logging
import os
import pandas as pandas
from openpyxl import load_workbook
from source.settings import (
    REPORT_TEMPLATE,
    REPORT_DIR
)


def generate_report(passwd_data, group_data):
    """
    Generate the report file
    """
    logging.info('Generating report ...')

    # report filename
    report_filename = 'dac-analysis-' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'

    # copy template report
    report_path = os.path.join(REPORT_DIR, report_filename)
    try:
        shutil.copy(REPORT_TEMPLATE, report_path)
    except OSError as err:
        logging.exception("Failed to copy the file %s - %s", report_filename, err.strerror)

    # write passwd information sheet
    data_frame = pandas.DataFrame(passwd_data, columns=['Username',
                                                        'Password',
                                                        'UID',
                                                        'GID',
                                                        'GECOS',
                                                        'Home Dir',
                                                        'Shell'])
    append_data_sheet(report_path, data_frame, sheet_name='passwd', header=0, index=False)

    # write group information sheet
    data_frame = pandas.DataFrame(group_data, columns=['Username',
                                                       'Password',
                                                       'GID',
                                                       'User List'])
    append_data_sheet(report_path, data_frame, sheet_name='group', header=0, index=False)

    logging.info('Generated report file %s', 'report/' + report_filename)



def append_data_sheet(filename, dataframe, sheet_name='New Sheet', startrow=None,
                      truncate_sheet=False, **to_excel_kwargs):
    """
    Append a dataframe to an existing Excel file

    @param filename: Excel path
    @type  filename: str

    @param sheet_name: Sheet name
    @type  sheet_name: str

    @param startrow: Start row
    @type  startrow: str

    @param truncate_sheet: Truncate sheet
    @type  truncate_sheet: bool

    @param to_excel_kwargs: dataframe.to_excel() arguments
    @type  to_excel_kwargs: list
    """

    if os.path.isfile(filename):
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pandas.ExcelWriter(
            filename, engine='openpyxl')  # pylint: disable=abstract-class-instantiated

        try:
            # read the report template
            writer.book = load_workbook(filename)

            # get the last row of the template
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate the sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)

            # copy sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except OSError as err:
            logging.exception("Failed to add data in the file %s | sheet name: %s | error: %s",
                              filename, sheet_name, err.errno)

        if startrow is None:
            startrow = 0

        # dataframe to excel
        dataframe.to_excel(writer, sheet_name,
                           startrow=startrow, **to_excel_kwargs)

        # save excel
        writer.save()

    else:
        logging.exception('Could not find the report file %s', filename)
